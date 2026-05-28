"""Run the full analytics pipeline and export dashboard-ready outputs."""

from __future__ import annotations

import argparse
import os
import shutil
import sys
import tempfile
from pathlib import Path

import pandas as pd

from .build_features import build_reporting_tables
from .clean_data import clean_all
from .config import (
    DOCS_IMAGES_DIR,
    PRIVATE_OUTPUTS_DIR,
    RAW_DATA_DIR,
    SAMPLE_DATA_DIR,
    SAMPLE_OUTPUTS_DIR,
    OutputPaths,
    ensure_directories,
    output_paths,
)
from .forecasting import build_forecast_results
from .kpis import build_working_capital_kpis, executive_summary_table
from .load_data import load_all_raw_data
from .utils import add_aging_bucket_order, save_csv
from .validation import validate_all


def _load_matplotlib():
    """Import matplotlib with a non-interactive backend."""

    try:
        cache_dir = Path(tempfile.gettempdir()) / "matapel_matplotlib_cache"
        cache_dir.mkdir(parents=True, exist_ok=True)
        os.environ.setdefault("MPLCONFIGDIR", str(cache_dir))
        os.environ.setdefault("XDG_CACHE_HOME", str(cache_dir))
        import matplotlib

        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError as exc:
        raise ImportError(
            "matplotlib is required for chart exports. Install dependencies with: pip install -r requirements.txt"
        ) from exc
    return plt


def _format_axis_currency(ax, label: str = "IDR", axis: str = "y") -> None:
    """Apply compact currency labels to a matplotlib axis."""

    from matplotlib.ticker import FuncFormatter

    def formatter(value: float, _position: int) -> str:
        abs_value = abs(value)
        if abs_value >= 1_000_000_000:
            return f"{value / 1_000_000_000:.1f}B {label}"
        if abs_value >= 1_000_000:
            return f"{value / 1_000_000:.1f}M {label}"
        if abs_value >= 1_000:
            return f"{value / 1_000:.1f}K {label}"
        return f"{value:,.0f} {label}"

    target_axis = ax.xaxis if axis == "x" else ax.yaxis
    target_axis.set_major_formatter(FuncFormatter(formatter))


def _save_empty_chart(path: Path, title: str, message: str) -> None:
    """Create a placeholder chart when a source table has no rows."""

    plt = _load_matplotlib()
    fig, ax = plt.subplots(figsize=(9, 5))
    ax.text(0.5, 0.5, message, ha="center", va="center", fontsize=12)
    ax.set_axis_off()
    fig.suptitle(title, fontsize=14, fontweight="bold")
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_monthly_sales(monthly_sales: pd.DataFrame, path: Path) -> None:
    """Export monthly revenue trend chart."""

    plt = _load_matplotlib()
    if monthly_sales.empty:
        _save_empty_chart(path, "Monthly Sales Trend", "No monthly sales data available.")
        return
    plot_df = monthly_sales.groupby("month", as_index=False)["revenue_idr"].sum().sort_values("month")
    fig, ax = plt.subplots(figsize=(11, 5.8))
    ax.plot(plot_df["month"], plot_df["revenue_idr"], marker="o", linewidth=2.2, color="#0f766e")
    ax.set_title("Monthly Sales Trend", fontsize=15, fontweight="bold", loc="left")
    ax.set_xlabel("Month")
    ax.set_ylabel("Revenue")
    _format_axis_currency(ax)
    ax.grid(axis="y", alpha=0.25)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_top_product_lines(product_line_performance: pd.DataFrame, path: Path) -> None:
    """Export top product line revenue chart."""

    plt = _load_matplotlib()
    if product_line_performance.empty:
        _save_empty_chart(path, "Top Product Lines by Revenue", "No product-line revenue data available.")
        return
    plot_df = (
        product_line_performance.groupby("product_line", as_index=False)["revenue_idr"]
        .sum()
        .sort_values("revenue_idr", ascending=False)
        .head(10)
        .sort_values("revenue_idr")
    )
    fig, ax = plt.subplots(figsize=(10.5, 6))
    ax.barh(plot_df["product_line"], plot_df["revenue_idr"], color="#2563eb")
    ax.set_title("Top Product Lines by Revenue", fontsize=15, fontweight="bold", loc="left")
    ax.set_xlabel("Revenue")
    _format_axis_currency(ax, axis="x")
    ax.grid(axis="x", alpha=0.2)
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_forecast(forecast_results: pd.DataFrame, path: Path) -> None:
    """Export company/product-line forecast chart."""

    plt = _load_matplotlib()
    if forecast_results.empty:
        _save_empty_chart(path, "Demand Forecast by Product Line", "No forecast data available.")
        return
    fig, ax = plt.subplots(figsize=(11, 6))
    for (company, product_line), subset in forecast_results.groupby(["company", "product_line"], dropna=False):
        subset = subset.sort_values("forecast_month")
        label = f"{company} | {product_line}"
        ax.plot(subset["forecast_month"], subset["forecast_quantity"], marker="o", linewidth=2, label=label[:55])
        ax.fill_between(subset["forecast_month"], subset["lower_bound"], subset["upper_bound"], alpha=0.12)
    ax.set_title("Next 3 Months Demand Forecast", fontsize=15, fontweight="bold", loc="left")
    ax.set_xlabel("Forecast Month")
    ax.set_ylabel("Forecast Quantity")
    ax.grid(axis="y", alpha=0.25)
    ax.legend(loc="best", fontsize=8)
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_aging_buckets(aging_summary: pd.DataFrame, path: Path, title: str) -> None:
    """Export AR/AP aging bucket chart."""

    plt = _load_matplotlib()
    if aging_summary.empty:
        _save_empty_chart(path, title, "No aging data available.")
        return
    plot_df = add_aging_bucket_order(aging_summary).groupby(
        ["aging_bucket_order", "aging_bucket"],
        as_index=False,
    )["outstanding"].sum()
    plot_df = plot_df.sort_values("aging_bucket_order")
    fig, ax = plt.subplots(figsize=(10, 5.8))
    ax.bar(plot_df["aging_bucket"].astype(str), plot_df["outstanding"], color="#9333ea")
    ax.set_title(title, fontsize=15, fontweight="bold", loc="left")
    ax.set_xlabel("Aging Bucket")
    ax.set_ylabel("Outstanding")
    _format_axis_currency(ax)
    ax.grid(axis="y", alpha=0.25)
    ax.tick_params(axis="x", rotation=25)
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_inventory_value(monthly_inventory_summary: pd.DataFrame, path: Path) -> None:
    """Export inventory value by product line chart."""

    plt = _load_matplotlib()
    if monthly_inventory_summary.empty:
        _save_empty_chart(path, "Inventory Value by Product Line", "No inventory data available.")
        return
    latest_month = monthly_inventory_summary["month"].max()
    plot_df = (
        monthly_inventory_summary[monthly_inventory_summary["month"].eq(latest_month)]
        .groupby("product_line", as_index=False)["inventory_value_idr"]
        .sum()
        .sort_values("inventory_value_idr", ascending=False)
        .head(10)
        .sort_values("inventory_value_idr")
    )
    fig, ax = plt.subplots(figsize=(10.5, 6))
    ax.barh(plot_df["product_line"], plot_df["inventory_value_idr"], color="#ea580c")
    ax.set_title("Inventory Value by Product Line", fontsize=15, fontweight="bold", loc="left")
    ax.set_xlabel("Inventory Value")
    _format_axis_currency(ax, axis="x")
    ax.grid(axis="x", alpha=0.2)
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def chart_working_capital(working_capital_kpis: pd.DataFrame, path: Path) -> None:
    """Export DSO/DPO/DIO/CCC working capital chart."""

    plt = _load_matplotlib()
    if working_capital_kpis.empty:
        _save_empty_chart(path, "Working Capital KPIs", "No working-capital data available.")
        return
    row = working_capital_kpis.iloc[0]
    plot_df = pd.DataFrame(
        {
            "metric": ["DSO", "DPO", "DIO", "CCC"],
            "days": [
                row["dso_days"],
                row["dpo_days"],
                row["dio_days"],
                row["cash_conversion_cycle_days"],
            ],
        }
    )
    fig, ax = plt.subplots(figsize=(8.5, 5.5))
    ax.bar(plot_df["metric"], plot_df["days"], color=["#0f766e", "#2563eb", "#ea580c", "#9333ea"])
    ax.set_title("Working Capital KPIs", fontsize=15, fontweight="bold", loc="left")
    ax.set_ylabel("Days")
    ax.grid(axis="y", alpha=0.25)
    fig.tight_layout()
    fig.savefig(path, dpi=160)
    plt.close(fig)


def export_charts(
    tables: dict[str, pd.DataFrame],
    forecast_results: pd.DataFrame,
    working_capital_kpis: pd.DataFrame,
    paths: OutputPaths,
) -> None:
    """Create all requested PNG chart outputs."""

    paths.charts.mkdir(parents=True, exist_ok=True)
    chart_monthly_sales(tables["monthly_sales_summary"], paths.charts / "monthly_sales_trend.png")
    chart_top_product_lines(tables["product_line_performance"], paths.charts / "top_product_lines_revenue.png")
    chart_forecast(forecast_results, paths.charts / "forecast_product_lines.png")
    chart_aging_buckets(tables["ar_aging_summary"], paths.charts / "ar_aging_buckets.png", "AR Aging Buckets")
    chart_aging_buckets(tables["ap_aging_summary"], paths.charts / "ap_aging_buckets.png", "AP Aging Buckets")
    chart_inventory_value(tables["monthly_inventory_summary"], paths.charts / "inventory_value_by_product_line.png")
    chart_working_capital(working_capital_kpis, paths.charts / "working_capital_kpis.png")


def copy_sample_charts_to_docs(paths: OutputPaths) -> None:
    """Copy sample-data chart PNGs into docs/images for the public README."""

    DOCS_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    for chart_path in paths.charts.glob("*.png"):
        shutil.copy2(chart_path, DOCS_IMAGES_DIR / chart_path.name)


def _clean_excel_value(value):
    """Convert NaN-like values for cleaner Excel output."""

    if pd.isna(value):
        return None
    return value


def _format_workbook(workbook) -> None:
    """Apply readable formatting to all Excel dashboard sheets."""

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="17324D")
    header_font = Font(color="FFFFFF", bold=True)

    date_headers = {"month", "forecast_month", "invoice_month", "due_month", "period_start", "period_end"}
    whole_number_headers = {
        "active_months",
        "counterparty_count",
        "customer_count",
        "delivery_metric_records",
        "document_count",
        "invoice_count",
        "invoice_line_count",
        "issue_count",
        "item_count",
        "late_delivery_count",
        "negative_inventory_lines",
        "negative_value_lines",
        "overdue_documents",
        "period_days",
        "po_lines",
        "product_line_count",
        "quantity_rank",
        "records_checked",
        "revenue_rank",
        "training_months",
        "zero_inventory_lines",
    }

    def is_date_header(header: str) -> bool:
        return header in date_headers or header.endswith("_date") or header.startswith("payment_date")

    def is_whole_number_header(header: str) -> bool:
        return header in whole_number_headers or header.endswith("_rank") or header.endswith("_count")

    def is_percent_header(header: str) -> bool:
        return header.endswith("_pct") or header in {"issue_rate", "gross_margin_pct"}

    for ws in workbook.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row and ws.max_column:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = str(ws.cell(row=1, column=col_idx).value or "").lower()
            sample_values = [ws.cell(row=row, column=col_idx).value for row in range(1, min(ws.max_row, 80) + 1)]
            max_len = max((len(str(_clean_excel_value(value))) for value in sample_values if _clean_excel_value(value) is not None), default=10)
            ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 34)

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value is None:
                    continue
                if is_date_header(header):
                    cell.number_format = "yyyy-mm-dd"
                elif is_percent_header(header):
                    cell.number_format = "0.0%"
                elif any(token in header for token in ["idr", "usd", "outstanding", "revenue", "cogs", "inventory_value", "amount"]):
                    cell.number_format = '#,##0;[Red](#,##0);"-"'
                elif is_whole_number_header(header):
                    cell.number_format = '#,##0;[Red](#,##0);"-"'
                elif any(token in header for token in ["qty", "quantity", "days", "forecast", "bound", "avg", "price"]):
                    cell.number_format = '#,##0.00;[Red](#,##0.00);"-"'
                cell.alignment = Alignment(vertical="top")


def export_excel_dashboard(
    tables: dict[str, pd.DataFrame],
    forecast_results: pd.DataFrame,
    working_capital_kpis: pd.DataFrame,
    data_quality_report: pd.DataFrame,
    paths: OutputPaths,
) -> Path:
    """Export a multi-sheet Excel workbook for dashboard consumption."""

    paths.excel.mkdir(parents=True, exist_ok=True)
    sheets = {
        "Executive_Summary": executive_summary_table(working_capital_kpis),
        "Sales_Monthly": tables["monthly_sales_summary"],
        "Product_Line_Performance": tables["product_line_performance"],
        "Forecast_Results": forecast_results,
        "Inventory_Summary": tables["monthly_inventory_summary"],
        "AR_Aging": tables["ar_aging_summary"],
        "AP_Aging": tables["ap_aging_summary"],
        "Working_Capital_KPIs": working_capital_kpis,
        "Data_Quality_Checks": data_quality_report,
    }
    with pd.ExcelWriter(paths.workbook, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        _format_workbook(writer.book)
    return paths.workbook


def export_tables(
    cleaned: dict[str, pd.DataFrame],
    tables: dict[str, pd.DataFrame],
    working_capital_kpis: pd.DataFrame,
    forecast_results: pd.DataFrame,
    data_quality_report: pd.DataFrame,
    paths: OutputPaths,
) -> None:
    """Export cleaned and reporting tables to CSV."""

    for name, df in cleaned.items():
        save_csv(df, paths.processed / f"{name}.csv")
    table_exports = dict(tables)
    table_exports["working_capital_kpis"] = working_capital_kpis
    table_exports["forecast_results"] = forecast_results
    table_exports["data_quality_report"] = data_quality_report
    for name, df in table_exports.items():
        save_csv(df, paths.tables / f"{name}.csv")


def run_pipeline(data_dir: Path = RAW_DATA_DIR, output_root: Path = PRIVATE_OUTPUTS_DIR) -> dict[str, Path]:
    """Run load, clean, validate, forecast, and export steps."""

    ensure_directories()
    paths = output_paths(output_root)
    raw = load_all_raw_data(data_dir=data_dir)
    cleaned = clean_all(raw)
    tables = build_reporting_tables(cleaned)
    working_capital_kpis = build_working_capital_kpis(
        cleaned["fact_sales"],
        cleaned["fact_inventory"],
        cleaned["fact_ar"],
        cleaned["fact_ap"],
    )
    forecast_results = build_forecast_results(tables["monthly_product_sales"])
    data_quality_report = validate_all(cleaned)

    export_tables(cleaned, tables, working_capital_kpis, forecast_results, data_quality_report, paths)
    export_charts(tables, forecast_results, working_capital_kpis, paths)
    export_excel_dashboard(tables, forecast_results, working_capital_kpis, data_quality_report, paths)

    return {
        "output_root": paths.root,
        "excel": paths.workbook,
        "forecast_results": paths.tables / "forecast_results.csv",
        "data_quality_report": paths.tables / "data_quality_report.csv",
        "monthly_sales_trend": paths.charts / "monthly_sales_trend.png",
    }


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Parse command line arguments."""

    parser = argparse.ArgumentParser(description="Run Matapel Chemicals operations analytics pipeline.")
    parser.add_argument(
        "--sample",
        action="store_true",
        help="Use anonymized/synthetic files from data/sample/ instead of confidential data/raw/ exports.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""

    args = parse_args(argv)
    data_dir = SAMPLE_DATA_DIR if args.sample else RAW_DATA_DIR
    output_root = SAMPLE_OUTPUTS_DIR if args.sample else PRIVATE_OUTPUTS_DIR
    try:
        outputs = run_pipeline(data_dir=data_dir, output_root=output_root)
    except Exception as exc:
        print(f"Pipeline failed: {exc}", file=sys.stderr)
        return 1

    print("Pipeline completed successfully.")
    for label, path in outputs.items():
        print(f"{label}: {path}")
    if args.sample:
        copy_sample_charts_to_docs(output_paths(output_root))
        print(f"sample_charts: {DOCS_IMAGES_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
